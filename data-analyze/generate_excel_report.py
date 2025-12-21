#!/usr/bin/env python3
"""
Generate Excel Report for Crasher Game Analysis
"""

import sqlite3
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelReportGenerator:
    """Generate comprehensive Excel report"""

    def __init__(self, db_path: str = "./crasher_data.db"):
        self.conn = sqlite3.connect(db_path)
        self.threshold = 2.0
        self.max_gap_minutes = 3

    def load_data(self) -> pd.DataFrame:
        """Load and prepare data"""
        query = "SELECT id, multiplier, bettor_count, timestamp FROM multipliers ORDER BY timestamp ASC"
        df = pd.read_sql_query(query, self.conn)
        df["timestamp"] = pd.to_datetime(df["timestamp"])

        # Detect sessions
        df["time_diff"] = df["timestamp"].diff()
        df["new_session"] = df["time_diff"] > timedelta(minutes=self.max_gap_minutes)
        df["session_id"] = df["new_session"].cumsum()
        if len(df) > 0:
            df.loc[0, "session_id"] = 0

        df["under_2x"] = df["multiplier"] < self.threshold

        return df

    def find_streaks(self, df: pd.DataFrame):
        """Find all streaks"""
        streaks = []
        current_streak = []

        for idx, row in df.iterrows():
            if row["under_2x"]:
                current_streak.append(
                    {
                        "id": row["id"],
                        "multiplier": row["multiplier"],
                        "timestamp": row["timestamp"],
                        "session_id": row["session_id"],
                    }
                )
            else:
                if current_streak:
                    streaks.append(
                        {
                            "length": len(current_streak),
                            "session_id": current_streak[0]["session_id"],
                            "start_time": current_streak[0]["timestamp"],
                            "end_time": current_streak[-1]["timestamp"],
                            "multipliers": [r["multiplier"] for r in current_streak],
                        }
                    )
                    current_streak = []

        if current_streak:
            streaks.append(
                {
                    "length": len(current_streak),
                    "session_id": current_streak[0]["session_id"],
                    "start_time": current_streak[0]["timestamp"],
                    "end_time": current_streak[-1]["timestamp"],
                    "multipliers": [r["multiplier"] for r in current_streak],
                }
            )

        return streaks

    def generate_excel(self, output_path: str):
        """Generate comprehensive Excel report"""
        print("Loading data...")
        df = self.load_data()
        streaks = self.find_streaks(df)

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Overview
        print("Creating Overview sheet...")
        self._create_overview_sheet(wb, df, streaks)

        # Sheet 2: Streak Analysis
        print("Creating Streak Analysis sheet...")
        self._create_streak_sheet(wb, streaks)

        # Sheet 3: Session Details
        print("Creating Session Details sheet...")
        self._create_session_sheet(wb, df)

        # Sheet 4: Risk Assessment
        print("Creating Risk Assessment sheet...")
        self._create_risk_sheet(wb, streaks)

        # Sheet 5: Raw Data
        print("Creating Raw Data sheet...")
        self._create_raw_data_sheet(wb, df)

        print(f"Saving to {output_path}...")
        wb.save(output_path)
        print("Excel report generated successfully!")

    def _create_overview_sheet(self, wb, df, streaks):
        """Create overview sheet"""
        ws = wb.create_sheet("Overview", 0)

        # Title
        ws["A1"] = "CRASHER GAME ANALYSIS REPORT"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:D1")

        # Summary stats
        ws["A3"] = "SUMMARY STATISTICS"
        ws["A3"].font = Font(size=12, bold=True)

        stats = [
            ["Metric", "Value"],
            ["Total Rounds", len(df)],
            ["Total Sessions", df["session_id"].nunique()],
            ["Date Range Start", df["timestamp"].min().strftime("%Y-%m-%d %H:%M")],
            ["Date Range End", df["timestamp"].max().strftime("%Y-%m-%d %H:%M")],
            ["", ""],
            ["Average Multiplier", f"=AVERAGE('Raw Data'!B:B)"],
            ["Max Multiplier", f"=MAX('Raw Data'!B:B)"],
            ["Min Multiplier", f"=MIN('Raw Data'!B:B)"],
            ["Rounds Under 2.0x", f"=COUNTIF('Raw Data'!E:E,TRUE)"],
            ["% Under 2.0x", f"=F11/B4"],
            ["", ""],
            [
                "Longest Streak (Under 2.0x)",
                max([s["length"] for s in streaks]) if streaks else 0,
            ],
            ["Total Streaks", len(streaks)],
            [
                "Average Streak Length",
                sum(s["length"] for s in streaks) / len(streaks) if streaks else 0,
            ],
        ]

        for row_idx, row_data in enumerate(stats, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:  # Header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

        # Format percentage
        ws["B14"].number_format = "0.0%"

        # Format numbers
        for row in range(7, 16):
            ws[f"B{row}"].number_format = "#,##0.00"

        # Column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 20

    def _create_streak_sheet(self, wb, streaks):
        """Create streak analysis sheet"""
        ws = wb.create_sheet("Streak Analysis")

        # Title
        ws["A1"] = "STREAK ANALYSIS (Consecutive Rounds Under 2.0x)"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:D1")

        # Distribution
        ws["A3"] = "Streak Distribution"
        ws["A3"].font = Font(size=12, bold=True)

        # Count streaks by length
        streak_counts = {}
        for streak in streaks:
            length = streak["length"]
            streak_counts[length] = streak_counts.get(length, 0) + 1

        headers = ["Streak Length", "Occurrences", "% of Total Streaks"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

        row_idx = 5
        total_streaks = len(streaks)
        for length in sorted(streak_counts.keys()):
            count = streak_counts[length]
            ws.cell(row=row_idx, column=1, value=length)
            ws.cell(row=row_idx, column=2, value=count)
            ws.cell(
                row=row_idx,
                column=3,
                value=count / total_streaks if total_streaks > 0 else 0,
            )
            ws[f"C{row_idx}"].number_format = "0.0%"

            # Highlight critical lengths
            if length >= 10:
                ws[f"A{row_idx}"].fill = PatternFill(
                    start_color="FFD700", fill_type="solid"
                )  # Gold
            if length >= 15:
                ws[f"A{row_idx}"].fill = PatternFill(
                    start_color="FF6B6B", fill_type="solid"
                )  # Red

            row_idx += 1

        # Specific counts (6-10)
        ws[f"A{row_idx + 2}"] = "CRITICAL STREAK COUNTS"
        ws[f"A{row_idx + 2}"].font = Font(size=12, bold=True)

        headers2 = ["Streak Length", "Count", "Notes"]
        for col_idx, header in enumerate(headers2, start=1):
            cell = ws.cell(row=row_idx + 3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

        critical_lengths = [
            (6, "Safe zone"),
            (7, "Safe zone"),
            (8, "Safe zone"),
            (9, "Safe zone"),
            (10, "Trigger point - Start betting"),
            (11, "1 loss in martingale"),
            (12, "2 losses in martingale"),
            (13, "3 losses in martingale"),
            (14, "4 losses in martingale"),
            (15, "5 losses - TOTAL WIPEOUT"),
        ]

        for i, (length, note) in enumerate(critical_lengths, start=row_idx + 4):
            ws.cell(row=i, column=1, value=length)
            ws.cell(row=i, column=2, value=streak_counts.get(length, 0))
            ws.cell(row=i, column=3, value=note)

            if length >= 15:
                ws[f"A{i}"].fill = PatternFill(start_color="FF6B6B", fill_type="solid")
                ws[f"B{i}"].fill = PatternFill(start_color="FF6B6B", fill_type="solid")

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 35

    def _create_session_sheet(self, wb, df):
        """Create session details sheet"""
        ws = wb.create_sheet("Session Details")

        # Calculate session stats
        session_stats = []
        for session_id in sorted(df["session_id"].unique()):
            session_df = df[df["session_id"] == session_id]
            session_stats.append(
                {
                    "Session ID": int(session_id),
                    "Total Rounds": len(session_df),
                    "Start Time": session_df["timestamp"].min(),
                    "End Time": session_df["timestamp"].max(),
                    "Duration (min)": (
                        session_df["timestamp"].max() - session_df["timestamp"].min()
                    ).total_seconds()
                    / 60,
                    "Avg Multiplier": session_df["multiplier"].mean(),
                    "Max Multiplier": session_df["multiplier"].max(),
                    "Rounds Under 2x": (session_df["multiplier"] < 2.0).sum(),
                    "% Under 2x": (session_df["multiplier"] < 2.0).sum()
                    / len(session_df)
                    * 100,
                }
            )

        session_df_final = pd.DataFrame(session_stats)

        # Write to sheet
        for r_idx, row in enumerate(
            dataframe_to_rows(session_df_final, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

        # Format
        for row in range(2, len(session_stats) + 2):
            ws[f"F{row}"].number_format = "0.00"
            ws[f"G{row}"].number_format = "0.00"
            ws[f"I{row}"].number_format = "0.0%"

        # Auto-width
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column].width = min(max_length + 2, 20)

    def _create_risk_sheet(self, wb, streaks):
        """Create risk assessment sheet"""
        ws = wb.create_sheet("Risk Assessment")

        # Title
        ws["A1"] = "RISK ASSESSMENT"
        ws["A1"].font = Font(size=14, bold=True)
        ws.merge_cells("A1:E1")

        # Strategy description
        ws["A3"] = "Strategy:"
        ws["B3"] = (
            "Wait for 10 consecutive rounds under 2.0x, then bet using martingale for up to 5 rounds"
        )
        ws["A3"].font = Font(bold=True)

        ws["A4"] = "Risk:"
        ws["B4"] = "15 consecutive rounds under 2.0x = Total bank wipeout"
        ws["A4"].font = Font(bold=True)
        ws["B4"].font = Font(color="FF0000")

        # Calculate outcomes
        trigger_streaks = [s for s in streaks if s["length"] >= 10]
        total_loss_count = sum(1 for s in streaks if s["length"] >= 15)
        partial_loss_count = sum(1 for s in streaks if 11 <= s["length"] <= 14)
        win_count = sum(1 for s in streaks if s["length"] == 10)

        # Summary
        ws["A7"] = "OUTCOME SUMMARY"
        ws["A7"].font = Font(size=12, bold=True)

        headers = ["Outcome", "Count", "Percentage", "Description"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

        outcomes = [
            ["WIN", win_count, f"=B9/B12", "Exactly 10 under 2x, then hit 2x+"],
            [
                "PARTIAL LOSS",
                partial_loss_count,
                f"=B10/B12",
                "11-14 under 2x, lost some rounds",
            ],
            [
                "TOTAL LOSS",
                total_loss_count,
                f"=B11/B12",
                "15+ under 2x, complete wipeout",
            ],
            [
                "TOTAL TRIGGERS",
                len(trigger_streaks),
                "100%",
                "Times strategy would activate",
            ],
        ]

        for row_idx, outcome in enumerate(outcomes, start=9):
            for col_idx, value in enumerate(outcome, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 12:  # Total row
                    cell.font = Font(bold=True)
                if row_idx == 11:  # Total loss
                    cell.fill = PatternFill(start_color="FF6B6B", fill_type="solid")

        # Format percentages
        for row in range(9, 12):
            ws[f"C{row}"].number_format = "0.0%"

        # Detailed breakdown
        ws["A15"] = "DETAILED STREAK BREAKDOWN"
        ws["A15"].font = Font(size=12, bold=True)

        breakdown_headers = ["Streak Length", "Session ID", "Start Time", "Multipliers"]
        for col_idx, header in enumerate(breakdown_headers, start=1):
            cell = ws.cell(row=16, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

        # Show longest streaks
        sorted_streaks = sorted(streaks, key=lambda x: x["length"], reverse=True)[:20]
        for row_idx, streak in enumerate(sorted_streaks, start=17):
            ws.cell(row=row_idx, column=1, value=streak["length"])
            ws.cell(row=row_idx, column=2, value=streak["session_id"])
            ws.cell(
                row=row_idx,
                column=3,
                value=streak["start_time"].strftime("%Y-%m-%d %H:%M"),
            )
            ws.cell(
                row=row_idx,
                column=4,
                value=", ".join([f"{m:.2f}" for m in streak["multipliers"][:10]])
                + ("..." if len(streak["multipliers"]) > 10 else ""),
            )

            if streak["length"] >= 15:
                ws[f"A{row_idx}"].fill = PatternFill(
                    start_color="FF6B6B", fill_type="solid"
                )

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 60

    def _create_raw_data_sheet(self, wb, df):
        """Create raw data sheet"""
        ws = wb.create_sheet("Raw Data")

        # Select columns
        export_df = df[
            ["id", "multiplier", "bettor_count", "timestamp", "under_2x", "session_id"]
        ].copy()
        export_df["timestamp"] = export_df["timestamp"].dt.strftime("%Y-%m-%d %H:%M:%S")

        # Write to sheet
        for r_idx, row in enumerate(
            dataframe_to_rows(export_df, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", fill_type="solid")

        # Format
        for row in range(2, len(df) + 2):
            ws[f"B{row}"].number_format = "0.00"

        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 25)

    def close(self):
        self.conn.close()


def main():
    import sys

    db_path = "./crasher_data.db" if len(sys.argv) < 2 else sys.argv[1]
    output_path = "data-analyze/outputs/crasher_analysis.xlsx"

    generator = ExcelReportGenerator(db_path)

    try:
        generator.generate_excel(output_path)
        print(f"\n✓ Excel report saved to: {output_path}")
    except Exception as e:
        print(f"Error: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
    finally:
        generator.close()


if __name__ == "__main__":
    main()
